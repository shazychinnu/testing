#!/usr/bin/env python3
"""
excel_vlookup_conn.py

Usage:
    python excel_vlookup_conn.py first_workbook.xlsx second_workbook.xlsx [output_workbook.xlsx]

- Copies the first sheet from second_workbook.xlsx into a new sheet named 'Conn' in first_workbook.xlsx.
- Adds a trailing column 'GS conn' to 'Conn' and fills it using VLOOKUP-style logic:
    =VLOOKUP(F2, 'CDR Summary By Investor'!B:D, 3, 0)
  Implemented in Python as an exact-match lookup.

Notes:
- The script finds the 'CDR Summary By Investor' sheet in the first workbook case-insensitively.
- The lookup uses positional columns B (key) and D (value) from that sheet (i.e., columns index 1 and 3).
- The Conn sheet's lookup key is taken from Excel column F (6th column). If the sheet has a column named 'F', that is used first; otherwise the 6th column header is used.
"""

import sys
import os
import pandas as pd
from openpyxl import load_workbook

def find_sheet_case_insensitive(sheet_names, target_name):
    target_lower = target_name.lower().replace(" ", "")
    for name in sheet_names:
        if name.lower().replace(" ", "") == target_lower:
            return name
    # fallback: if any name contains the words ignoring spaces/case
    for name in sheet_names:
        if target_lower in name.lower().replace(" ", ""):
            return name
    return None

def main(file1, file2, output_file=None):
    if output_file is None:
        # default to overwrite input file1
        output_file = file1

    if not os.path.exists(file1):
        print(f"Error: file1 '{file1}' not found.")
        return
    if not os.path.exists(file2):
        print(f"Error: file2 '{file2}' not found.")
        return

    # Read sheet names from first workbook
    with pd.ExcelFile(file1, engine="openpyxl") as xls1:
        sheet_names_1 = xls1.sheet_names

    # Find the CDR Summary By Investor sheet (case-insensitive)
    target_sheet_name = find_sheet_case_insensitive(sheet_names_1, "CDR Summary By Investor")
    if target_sheet_name is None:
        print("Error: Could not find a sheet named like 'CDR Summary By Investor' in the first workbook.")
        print("Available sheets:", sheet_names_1)
        return

    # Read the lookup sheet (from file1)
    lookup_df = pd.read_excel(file1, sheet_name=target_sheet_name, engine="openpyxl", dtype=object)
    if lookup_df.shape[1] < 4:
        print(f"Error: The lookup sheet '{target_sheet_name}' must have at least 4 columns (B:D present).")
        print("Found columns:", list(lookup_df.columns))
        return

    # Read the first sheet of file2 (source to copy into Conn)
    with pd.ExcelFile(file2, engine="openpyxl") as xls2:
        sheet_names_2 = xls2.sheet_names
        if len(sheet_names_2) == 0:
            print("Error: second workbook contains no sheets.")
            return
        source_sheet_name = sheet_names_2[0]
    conn_source_df = pd.read_excel(file2, sheet_name=0, engine="openpyxl", dtype=object)

    # Make a copy for the Conn sheet
    conn_df = conn_source_df.copy()

    # Determine the Conn lookup column (Excel column F => 6th column)
    # Priority: if there's a column literally named 'F' (case-sensitive), use it.
    if 'F' in conn_df.columns:
        conn_lookup_col = 'F'
    else:
        if conn_df.shape[1] >= 6:
            conn_lookup_col = conn_df.columns[5]  # 0-indexed; 5 => 6th column => Excel F
        else:
            print("Error: Conn sheet (from second workbook) has fewer than 6 columns and no column named 'F'.")
            print("Columns found:", list(conn_df.columns))
            return

    # Build mapping from lookup sheet: key from column B (index 1), value from column D (index 3)
    # Convert keys to str and strip whitespace to emulate Excel exact-match behavior more robustly.
    lookup_keys = lookup_df.iloc[:, 1].astype(str).str.strip().fillna('')
    lookup_values = lookup_df.iloc[:, 3]  # keep original type for returned values
    mapping = dict(zip(lookup_keys, lookup_values))

    # Prepare Conn lookup series (as strings stripped)
    conn_keys_series = conn_df[conn_lookup_col].astype(str).str.strip().fillna('')

    # Perform exact-match lookup (like VLOOKUP with range B:D and col_index 3)
    # If no match, result will be NaN
    conn_df['GS conn'] = conn_keys_series.map(mapping)

    # Save the updated workbook: load original with openpyxl, then write Conn sheet (replace if exists)
    # We'll use pandas ExcelWriter with engine openpyxl and pass the loaded workbook as the book.
    from openpyxl import load_workbook
    wb = load_workbook(filename=file1)

    # Remove existing 'Conn' sheet if present
    if 'Conn' in wb.sheetnames:
        std = wb['Conn']
        wb.remove(std)

    # Use ExcelWriter with the existing workbook
    with pd.ExcelWriter(output_file, engine="openpyxl", mode="a" if output_file == file1 else "w") as writer:
        # If mode='a' we need to set book to existing workbook
        # But pandas openpyxl engine will attempt to open file; to be robust, save workbook first if same file
        if output_file == file1:
            writer.book = wb
            # ensure writer.sheets mapping exists for preservation
            writer.sheets = {ws.title: ws for ws in wb.worksheets}
        # Write the Conn sheet at the end
        conn_df.to_excel(writer, sheet_name='Conn', index=False)
        # If we created a new file (output_file != file1), we also want to preserve other sheets from original workbook.
        # For simplicity: if writing to a different output file, copy all sheets from original into writer, then overwrite Conn.
        if output_file != file1:
            # load original and write all sheets (except if overloaded by conn_df) - handled below
            original_wb = load_workbook(filename=file1)
            for ws in original_wb.worksheets:
                name = ws.title
                if name == 'Conn':
                    continue
                # read sheet into DataFrame (safe way)
                df_sheet = pd.read_excel(file1, sheet_name=name, engine="openpyxl", dtype=object)
                df_sheet.to_excel(writer, sheet_name=name, index=False)

    print(f"Success. 'Conn' sheet added/updated in '{output_file}'.")
    print(f" - Source Conn data taken from '{file2}' sheet '{source_sheet_name}'.")
    print(f" - Lookup used sheet '{target_sheet_name}' from '{file1}', columns B (key) and D (value).")
    print(" - The new column 'GS conn' has been appended to the Conn sheet.")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python excel_vlookup_conn.py first_workbook.xlsx second_workbook.xlsx [output_workbook.xlsx]")
    else:
        file1 = sys.argv[1]
        file2 = sys.argv[2]
        out = sys.argv[3] if len(sys.argv) >= 4 else None
        main(file1, file2, out)
