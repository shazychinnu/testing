import re
import pandas as pd
from openpyxl import load_workbook, Workbook

cdr_file = "CDR_VREP.xlsx"
wizard_file = "report_file.xlsx"
output_file = "output.xlsx"

# -----------------------------
# Utilities
# -----------------------------
def ensure_output_file_exists():
    try:
        load_workbook(output_file)
    except FileNotFoundError:
        Workbook().save(output_file)

def delete_sheet_if_exists(path, sheet_name):
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        wb.save(path)
    wb.close()

def norm_key(x) -> str:
    """Normalize keys so joins are reliable."""
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = s.replace("\u00A0", " ")
    s = re.sub(r"[ ,\-]", "", s)
    return s.upper()

# -----------------------------
# Step 1: Commitment Sheet
# -----------------------------
def create_commitment_sheet():
    ensure_output_file_exists()
    delete_sheet_if_exists(output_file, "Commitment Sheet")

    # ---- CDR Summary (GS source) ----
    cdr = pd.read_excel(
        cdr_file,
        sheet_name="CDR Summary By Investor",
        engine="openpyxl",
        skiprows=2,
    )
    cdr.columns = cdr.columns.str.strip()
    cdr["Account Number"] = cdr["Account Number"].apply(norm_key)

    # robust numeric source
    cdr["Investor Commitment"] = pd.to_numeric(cdr["Investor Commitment"], errors="coerce")
    acctnorm_to_commitment = cdr.set_index("Account Number")["Investor Commitment"].to_dict()

    # ---- Data_format (left) ----
    df = pd.read_excel(wizard_file, sheet_name="Data_format", engine="openpyxl")
    df.columns = df.columns.str.strip()

    required = ["Legal Entity", "Bin ID", "Investran Acct ID", "Commitment Amount"]
    for col in required:
        if col not in df.columns:
            raise KeyError(f"Missing required column in Data_format: '{col}'")

    df["Legal Entity"] = df["Legal Entity"].astype(str).str.strip()
    df["Commitment Amount"] = pd.to_numeric(df["Commitment Amount"], errors="coerce").fillna(0)

    df["_bin_norm"] = df["Bin ID"].apply(norm_key)
    df["_inv_acct_norm"] = df["Investran Acct ID"].apply(norm_key)

    subtotal_mask = df["Legal Entity"].str.contains("Subtotal", case=False, na=False)

    # ---- GS Commitment (Bin ID ↔ Account Number), only where Bin ID present & not subtotal ----
    df["GS Commitment"] = pd.NA
    valid_gs = (~subtotal_mask) & df["_bin_norm"].ne("") & df["_bin_norm"].notna()
    df.loc[valid_gs, "GS Commitment"] = df.loc[valid_gs, "_bin_norm"].map(acctnorm_to_commitment)
    df["GS Commitment"] = pd.to_numeric(df["GS Commitment"], errors="coerce")

    # Subtotals per block (above each subtotal row)
    sub_idxs = df.index[subtotal_mask].tolist()
    if sub_idxs:
        sub_idxs.append(len(df))
        start = 0
        for idx in sub_idxs[:-1]:
            block = df.iloc[start:idx]
            df.at[idx, "Commitment Amount"] = pd.to_numeric(block["Commitment Amount"], errors="coerce").sum(skipna=True)
            df.at[idx, "GS Commitment"] = pd.to_numeric(block["GS Commitment"], errors="coerce").sum(skipna=True)
            start = idx + 1

    # Per rule: blank GS on subtotal rows and compute GS Check
    df.loc[subtotal_mask, "GS Commitment"] = pd.NA
    df["GS Check"] = pd.to_numeric(df["Commitment Amount"], errors="coerce") - pd.to_numeric(df["GS Commitment"], errors="coerce")

    # ---- SS Commitment (Investran Acct ID ↔ Investor ID) ----
    # Build SS source: sum of Commitment Amount by normalized Investran Acct ID (non-blank only)
    ss_source = (
        df.loc[df["_inv_acct_norm"].ne("") & df["_inv_acct_norm"].notna()]
        .groupby("_inv_acct_norm", as_index=True)["Commitment Amount"]
        .sum()
        .to_dict()
    )

    investern = pd.read_excel(wizard_file, sheet_name="investern_format", engine="openpyxl")
    investern.columns = investern.columns.str.strip()

    for col in ["Investor ID", "Invester Commitment"]:
        if col not in investern.columns:
            raise KeyError(f"Missing required column in investern_format: '{col}'")

    investern["Investor ID"] = investern["Investor ID"].astype(str).str.strip().str.upper()
    investern["_id_norm"] = investern["Investor ID"].apply(norm_key)
    investern["Invester Commitment"] = pd.to_numeric(investern["Invester Commitment"], errors="coerce")

    # Map SS strictly by normalized key; leave others NaN
    investern["SS Commitment"] = investern["_id_norm"].map(ss_source)
    investern["SS Commitment"] = pd.to_numeric(investern["SS Commitment"], errors="coerce")
    investern["SS Check"] = investern["SS Commitment"] - investern["Invester Commitment"]

    # ---- Combine left + spacer + right (no totals yet) ----
    max_rows = max(len(df), len(investern))
    spacer = pd.DataFrame({f"Empty_{i}": [""] * max_rows for i in range(3)})
    left = df.reindex(range(max_rows)).reset_index(drop=True)
    right = investern.reindex(range(max_rows)).reset_index(drop=True)
    combined_df = pd.concat([left, spacer, right], axis=1)

    # ---- Append exactly one SS subtotal row at bottom (calculate totals from 'investern', not combined_df) ----
    ss_total_commit = pd.to_numeric(investern["SS Commitment"], errors="coerce").sum(skipna=True)
    ss_total_invest = pd.to_numeric(investern["Invester Commitment"], errors="coerce").sum(skipna=True)
    ss_total_check = ss_total_commit - ss_total_invest

    ss_total_row = {col: "" for col in combined_df.columns}
    ss_total_row.update({
        "Vehicle/Investor": "Subtotal (SS Total)",
        "Investor ID": "",
        "Invester Commitment": ss_total_invest,
        "SS Commitment": ss_total_commit,
        "SS Check": ss_total_check,
    })
    combined_df = pd.concat([combined_df, pd.DataFrame([ss_total_row])], ignore_index=True)

    # ---- Blank SS where Investor ID is missing (after combine, before export) ----
    if "Investor ID" in combined_df.columns:
        inv_id_norm = combined_df["Investor ID"].astype(str).str.strip().str.upper()
        mask_blank_ss = inv_id_norm.isna() | inv_id_norm.eq("") | inv_id_norm.isin(["NAN", "NONE", "NULL"])
        for col in ["SS Commitment", "SS Check", "Invester Commitment"]:
            if col in combined_df.columns:
                combined_df.loc[mask_blank_ss, col] = pd.NA

    # ---- Make sure numeric columns are numeric (prevents float+str errors) ----
    numeric_cols = [
        "GS Commitment", "Commitment Amount", "GS Check",
        "SS Commitment", "SS Check", "Invester Commitment"
    ]
    for col in numeric_cols:
        if col in combined_df.columns:
            combined_df[col] = pd.to_numeric(combined_df[col], errors="coerce")

    # ---- Final cosmetic: replace NaNs with blanks for Excel only ----
    combined_df = combined_df.fillna("")

    with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as w:
        combined_df.to_excel(w, sheet_name="Commitment Sheet", index=False)

    print("✅ Commitment Sheet created (clean, no NaN, no type-mix).")
    return combined_df

# -----------------------------
# Step 2: Entry Sheet
# -----------------------------
def create_entry_sheet_with_subtotals(commitment_df):
    delete_sheet_if_exists(output_file, "Entry")

    df_raw = pd.read_excel(wizard_file, sheet_name="allocation_data", engine="openpyxl", header=None)
    header_rows = df_raw.index[df_raw.iloc[:, 0].astype(str) == "Vehicle/Investor"].tolist()
    tables = []

    for i, h in enumerate(header_rows):
        start = h
        end = header_rows[i + 1] if i + 1 < len(header_rows) else len(df_raw)
        block = df_raw.iloc[start:end].reset_index(drop=True)
        block.columns = block.iloc[0]
        block = block.drop(0).reset_index(drop=True)

        if "Final LE Amount" in block.columns:
            block["Final LE Amount"] = pd.to_numeric(block["Final LE Amount"], errors="coerce")
            subtotal_val = block["Final LE Amount"].sum(skipna=True)
        else:
            subtotal_val = 0

        subtotal_row = {col: "" for col in block.columns}
        subtotal_row[block.columns[0]] = "Subtotal"
        if "Final LE Amount" in block.columns:
            subtotal_row["Final LE Amount"] = subtotal_val

        block = pd.concat([block, pd.DataFrame([subtotal_row])], ignore_index=True)
        tables.append(block)

    final_df = pd.concat(tables, ignore_index=True) if tables else pd.DataFrame()

    # Normalize ID for mapping
    id_col = "Investor ID" if "Investor ID" in final_df.columns else "Investor Id"
    final_df["_id_norm"] = final_df[id_col].apply(norm_key)

    # Build maps from commitment_df (left side data)
    cm = commitment_df.copy()
    cm["_inv_acct_norm"] = cm["Investran Acct ID"].apply(norm_key)

    id_to_bin = (
        cm.dropna(subset=["Bin ID"])
          .drop_duplicates(subset=["_inv_acct_norm"])
          .set_index("_inv_acct_norm")["Bin ID"]
          .to_dict()
    )
    id_to_amt = cm.groupby("_inv_acct_norm", as_index=True)["Commitment Amount"].sum().to_dict()

    final_df["Bin ID"] = final_df["_id_norm"].map(id_to_bin)
    final_df["Commitment Amount"] = final_df["_id_norm"].map(id_to_amt)

    # Final cosmetic cleanup for Entry sheet
    final_df = final_df.fillna("")

    with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as w:
        final_df.to_excel(w, sheet_name="Entry", index=False)

    print("✅ Entry Sheet created (clean).")

# -----------------------------
# Main
# -----------------------------
if __name__ == "__main__":
    commitment_df = create_commitment_sheet()
    create_entry_sheet_with_subtotals(commitment_df)
    print("🎯 Automation completed successfully!")
