import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook

cdr_file = "CDR_File.xlsx"
wizard_file = "Wizard_File.xlsx"
output_file = "Output_File.xlsx"

# =====================================================================
# BASIC UTILITIES
# =====================================================================

cdr_file_data = pd.read_excel(
    cdr_file,
    sheet_name="CDR Summary By Investor",
    engine="openpyxl",
    skiprows=2
)

def find_col(df, target):
    """Find a column ignoring case and spaces."""
    target_clean = str(target).strip().lower().replace(" ", "")
    for col in df.columns:
        if pd.isna(col):
            continue
        col_clean = str(col).strip().lower().replace(" ", "")
        if col_clean == target_clean:
            return col
    return None

def ensure_output_file_exists():
    try:
        with open(output_file, "rb"):
            pass
    except FileNotFoundError:
        wb = load_workbook(cdr_file)
        wb.save(output_file)

def delete_sheet_if_exists(file_path, sheet_name):
    try:
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)
            wb.save(file_path)
    except FileNotFoundError:
        pass

def norm_key(x):
    """Normalize ID or Bin ID."""
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s.endswith(".."):
        s = s[:-2]
    s = s.replace("\u00A0", " ")
    s = re.sub(r"[,\-]", "", s)
    return s.upper()

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.replace(
        [pd.NA, None, np.nan, "NaN", "<NA>", "None", "NULL", "nan"], ""
    )
    return df.astype(object)

# =====================================================================
# COMMITMENT SHEET (UPDATED WITH MULTI-KEY MATCHING)
# =====================================================================

def create_commitment_sheet():
    ensure_output_file_exists()
    delete_sheet_if_exists(output_file, "Commitment Sheet")

    # 1. Load CDR
    cdr = cdr_file_data.copy()
    cdr.columns = cdr.columns.str.strip()
    cdr["Account Number"] = cdr["Account Number"].apply(norm_key)
    cdr["Investor Commitment"] = pd.to_numeric(
        cdr["Investor Commitment"], errors="coerce"
    ).fillna(0)

    # BUILD MULTI-KEY MATCH: (Bin ID, Account Number)
    cdr["_bin_norm"] = cdr["Account Number"].apply(norm_key)
    cdr["_acct_norm"] = cdr["Account Number"].apply(norm_key)

    acct_to_commit_multi = (
        cdr.set_index(["_bin_norm", "_acct_norm"])["Investor Commitment"].to_dict()
    )

    # 2. Load Data_format
    df = pd.read_excel(wizard_file, sheet_name="Data_format", engine="openpyxl")
    df.columns = df.columns.str.strip()
    df["Legal Entity"] = df["Legal Entity"].astype(str).str.strip()
    df["Commitment Amount"] = pd.to_numeric(df["Commitment Amount"], errors="coerce").fillna(0)

    df["_bin_norm"] = df["Bin ID"].apply(norm_key)
    df["_inv_acct_norm"] = df["Investran Acct ID"].apply(norm_key)

    # APPLY MULTI-KEY GS COMMITMENT LOOKUP
    def lookup_gs(row):
        key = (row["_bin_norm"], row["_inv_acct_norm"])
        return acct_to_commit_multi.get(key, 0)

    df["GS Commitment"] = df.apply(lookup_gs, axis=1)
    df["GS Commitment"] = pd.to_numeric(df["GS Commitment"], errors="coerce").fillna(0)

    # Replace Commitment Amount when GS available
    df.loc[
        (df["Commitment Amount"] == 0) & (df["GS Commitment"] != 0),
        "Commitment Amount"
    ] = df["GS Commitment"]

    # SUBTOTAL CALCULATION
    subtotal_mask = df["Legal Entity"].str.contains("Subtotal", case=False, na=False)
    subtotal_indices = df.index[subtotal_mask].tolist()

    start_idx = 0
    for idx in subtotal_indices:
        section = df.iloc[start_idx:idx]
        total_commit = section["Commitment Amount"].sum()
        total_gs_commit = section["GS Commitment"].sum()

        df.at[idx, "Commitment Amount"] = total_commit
        df.at[idx, "GS Commitment"] = total_gs_commit
        df.at[idx, "GS Check"] = total_commit - total_gs_commit

        start_idx = idx + 1

    df["GS Check"] = df["Commitment Amount"] - df["GS Commitment"]

    # 4. SS Commitment
    ss_source = (
        df[df["_bin_norm"] != ""]
        .groupby("_bin_norm")["Commitment Amount"]
        .sum()
        .to_dict()
    )

    investern = pd.read_excel(
        wizard_file, sheet_name="investern_format", engine="openpyxl"
    )
    investern.columns = investern.columns.str.strip()
    investern["Account Number"] = (
        investern["Account Number"]
        .astype(str)
        .str.strip()
        .str.upper()
        .replace(["NAN", "NONE", "NULL", "PD.NA", "<NA>", "NA", "N/A"], "")
    )
    investern["Account Number"] = investern["Account Number"].where(
        investern["Account Number"] != "nan", ""
    )

    investern["_id_norm"] = investern["Account Number"].apply(norm_key)
    investern["Invester Commitment"] = pd.to_numeric(
        investern["Invester Commitment"], errors="coerce"
    ).fillna(0)

    investern["SS Commitment"] = pd.to_numeric(
        investern["_id_norm"].map(ss_source), errors="coerce"
    ).fillna(0)

    investern["SS Check"] = (
        investern["SS Commitment"] - investern["Invester Commitment"]
    )

    # 5. Combine
    max_rows = max(len(df), len(investern))
    spacer = pd.DataFrame(
        {f"Empty_{i}": [""] * max_rows for i in range(3)}, dtype=object
    )

    df = df.reindex(range(max_rows)).reset_index(drop=True)
    investern = investern.reindex(range(max_rows)).reset_index(drop=True)

    combined_df = pd.concat(
        [df.astype(object), spacer, investern.astype(object)], axis=1
    )

    # 6. Add SS subtotal
    ss_total_commit = investern["SS Commitment"].sum()
    ss_total_invest = investern["Invester Commitment"].sum()
    ss_total_check = ss_total_commit - ss_total_invest

    subtotal_row = {col: "" for col in combined_df.columns}
    subtotal_row.update({
        "Vehicle/Investor": "Subtotal (SS Total)",
        "Invester Commitment": ss_total_invest,
        "SS Commitment": ss_total_commit,
        "SS Check": ss_total_check,
    })

    combined_df = pd.concat([combined_df, pd.DataFrame([subtotal_row])], ignore_index=True)

    # 7. Remove helper cols
    helper_cols = [col for col in combined_df.columns if col.startswith("_")]
    combined_df.drop(columns=helper_cols, errors="ignore", inplace=True)

    # 8. Final cleanup
    combined_df = clean_dataframe(combined_df)

    # 9. Write
    with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:
        combined_df.to_excel(writer, sheet_name="Commitment Sheet", index=False)

    print("Commitment Sheet created successfully.")
    return combined_df

# =====================================================================
# CLEAN HEADERS
# =====================================================================

def clean_headers(headers):
    cleaned = []
    seen = {}
    for i, col in enumerate(headers):
        col_str = str(col).strip() if pd.notna(col) and str(col).strip() else f"Unnamed_{i}"
        if col_str in seen:
            seen[col_str] += 1
            col_str = f"{col_str}_{seen[col_str]}"
        else:
            seen[col_str] = 0
        cleaned.append(col_str)
    return cleaned

# =====================================================================
# REMOVE EMPTY OR ZERO COLUMNS
# =====================================================================

def remove_empty_or_zero_columns(df):
    cleaned = df.copy()

    if "_id_norm" in cleaned.columns:
        cleaned = cleaned.drop(columns=["_id_norm", "ExternalExpenses__bin_id_form"], errors="ignore")

    remove_total_cols = []
    for col in cleaned.columns:
        col_str = str(col).strip()
        if col_str.lower() == "total":
            remove_total_cols.append(col)
        elif re.match(r".*_Total(\.\d+)?$", col_str, flags=re.IGNORECASE):
            remove_total_cols.append(col)

    cleaned.drop(columns=remove_total_cols, errors="ignore", inplace=True)

    cols_to_drop = []
    for col in cleaned.columns:
        if col == cleaned.columns[0]:
            continue

        s = cleaned[col].replace("", pd.NA)
        if not pd.to_numeric(s, errors="coerce").notna().any():
            if s.isna().all():
                cols_to_drop.append(col)
            continue

        s_num = pd.to_numeric(s, errors="coerce")
        if s.isna().all() or s_num.fillna(0).eq(0).all():
            cols_to_drop.append(col)

    return cleaned.drop(columns=cols_to_drop, errors="ignore")

# =====================================================================
# ENTRY SHEET (UNCHANGED — ORIGINAL LOGIC)
# =====================================================================

def create_entry_sheet_with_subtotals(commitment_df):
    delete_sheet_if_exists(output_file, "Entry")

    df_raw = pd.read_excel(
        wizard_file, sheet_name="allocation_data", engine="openpyxl", header=None
    )

    header_rows = df_raw.index[
        df_raw.iloc[:, 0].astype(str) == "Vehicle/Investor"
    ].tolist()

    cdr_summary = pd.read_excel(
        cdr_file,
        sheet_name="CDR Summary By Investor",
        engine="openpyxl",
        skiprows=2
    )

    if header_rows:
        header_values = clean_headers(list(df_raw.loc[header_rows[0]]))
        df_raw.columns = header_values
        df_raw = df_raw.drop(header_rows[0]).reset_index(drop=True)

    id_col = find_col(df_raw, "Investor ID") or find_col(df_raw, "Investor Id")
    if not id_col:
        raise KeyError("Investor ID column missing in allocation data")

    df_raw["_id_norm"] = df_raw[id_col].apply(norm_key)

    cm = commitment_df.copy()
    cm["_inv_acct_norm"] = cm["Investran Acct ID"].apply(norm_key)

    id_to_bin = (
        cm.dropna(subset=["Bin ID"])
        .drop_duplicates(subset=["_inv_acct_norm"])
        .set_index("_inv_acct_norm")["Bin ID"]
        .to_dict()
    )

    cdr_summary["Account Number"] = cdr_summary["Account Number"].astype(str).str.upper()
    cdr_summary["_bin_id_form"] = cdr_summary["Account Number"].apply(norm_key)

    id_to_amt = (
        cdr_summary.groupby("_bin_id_form")["Investor Commitment"]
        .sum()
        .to_dict()
    )

    df_raw["Bin ID"] = df_raw["_id_norm"].map(id_to_bin)

    df_raw["Commitment Amount"] = df_raw["Bin ID"].apply(
        lambda x: id_to_amt.get(norm_key(x), "") if pd.notna(x) and str(x).strip() else ""
    )

    cdr_summary.columns = clean_headers(cdr_summary.columns)

    new_columns = []
    section_cols = []
    section = None

    for col in cdr_summary.columns:
        col_clean = str(col).lower().strip()
        if col_clean == "total contributions to commitment":
            section = "Contributions"
        elif col_clean == "total recallable":
            section = "Distributions"
        elif col_clean == "external expenses":
            section = "ExternalExpenses"

        if section and col not in ["Investor ID", "Account Number", "Investor Name", "Bin ID"]:
            new_col = f"{section}_{col}"
            section_cols.append(new_col)
            new_columns.append(new_col)
        else:
            new_columns.append(col)

    cdr_summary.columns = new_columns

    cdr_bin_col = find_col(cdr_summary, "Account Number")
    cdr_summary[cdr_bin_col] = cdr_summary[cdr_bin_col].astype(str).apply(norm_key)
    cdr_summary = cdr_summary.drop_duplicates(subset=[cdr_bin_col])
    cdr_indexed = cdr_summary.set_index(cdr_bin_col)

    tables = []

    for i, h in enumerate(header_rows):
        start = h
        end = header_rows[i+1] if (i+1) < len(header_rows) else len(df_raw)

        block = df_raw.loc[start:end].reset_index(drop=True)
        block = block.drop(0).reset_index(drop=True)
        block.columns = df_raw.columns

        entry_bin_col = find_col(block, "Bin ID")

        for col in section_cols:
            if entry_bin_col and col in cdr_indexed.columns:
                block[col] = block[entry_bin_col].apply(
                    lambda x: cdr_indexed[col].get(norm_key(x), "")
                    if pd.notna(x) and str(x).strip() else ""
                )
            else:
                block[col] = ""

        numeric_like = []
        for col in block.columns:
            s = block[col].replace("", pd.NA)
            if pd.to_numeric(s, errors="coerce").notna().any():
                numeric_like.append(col)

        subtotal_row = {col: "" for col in block.columns}
        for col in numeric_like:
            if col in ["Investor ID", "_id_norm"]:
                continue
            subtotal_row[col] = pd.to_numeric(block[col], errors="coerce").sum()

        subtotal_row[block.columns[0]] = "Subtotal"
        block = pd.concat([block, pd.DataFrame([subtotal_row])], ignore_index=True)

        tables.append(block)

    final_df = pd.concat(tables, ignore_index=True)

    final_df = remove_empty_or_zero_columns(final_df)
    final_df = clean_dataframe(final_df)

    with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:
        final_df.to_excel(writer, sheet_name="Entry", index=False)

    print("Entry Sheet created successfully.")

# =====================================================================
# MAIN
# =====================================================================

if __name__ == "__main__":
    commitment_df = create_commitment_sheet()
    create_entry_sheet_with_subtotals(commitment_df)
    print("Automation completed successfully.")
